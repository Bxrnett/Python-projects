import win32com.client
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd

# Opens the workbook
path = 'file path here'
wb = load_workbook(path + 'changewindow.xlsx')
ws = wb.active

# Connect to Outlook
outlook = win32com.client.Dispatch("Outlook.Application").GetNameSpace("MAPI")
folder_name = "Monitor Change Window"  # Specific folder where emails are grabbed from
folder = outlook.GetDefaultFolder(6) 

# Get the most recent email in the specified folder
email = folder.Items.GetLast()

# This will write the email contents to the spreadsheet in the specific areas
ws.cell(row=8, column=5).value = email.body

# This is where the excel spreadsheet will be saved, as it is getting overwrited daily, it is saved to the same file path
wb.save("file path here") 

# This line below will read the excell data
df = pd.read_excel('file path here')

# Write the information collected to an HTML file
df.to_html('changewindow.html', index=False)


ws = load_workbook('file path here').active
email = win32com.client.Dispatch("Outlook.Application").GetNameSpace("MAPI").GetDefaultFolder(6).Items.GetLast()
ws.cell(row=8, column=5).value = email.body
wb.save("file path here")